# Utility function
# Adds a row or a col of values to an Excel xlsx file

import argparse
import logging
import init_utils as iu
import import_utils as util

import os
import os.path as op
_mydir = op.abspath(op.dirname(__file__))
_djangodir = op.normpath(op.join(_mydir, '../django'))
import sys
sys.path.insert(0, _djangodir)
import chdir as cd
with cd.chdir(_djangodir):
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'hmslincs_server.settings')
    import django.db.models as models
    import django.db.models.fields as fields
del _mydir, _djangodir

__version__ = "$Revision: 24d02504e664 $"
# $Source$

# ---------------------------------------------------------------------------

import setparams as _sg
_params = dict(
    VERBOSE = False,
    APPNAME = 'db',
)
_sg.setparams(_params)
del _sg, _params

# ---------------------------------------------------------------------------

logger = logging.getLogger(__name__)
from django.utils.encoding import smart_str
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet import Worksheet

def new_row(path,newpath,val1,val2,
            sheet_name=None,
            row1=None, col1=None):
    logger.info(str(("new row for sheet",sheet_name, val1,val2)))

    
    wb = load_workbook(path)    
    sheet = wb.get_sheet_by_name(sheet_name);

    num_rows = sheet.get_highest_row()
    start_col = col1 or 0
    start_row = row1 or num_rows
    
    cell = sheet.cell(row = start_row, column = start_col)
    cell.value = val1
    cell = sheet.cell(row = start_row, column = start_col+1)
    cell.value = val2
    
    wb.save(newpath)   
    print 'wrote new workbook row:', newpath, val1, val2

def new_col(path,newpath,val1,val2,
            sheet_name=None,
            row1=None, col1=None):
    logger.info(str(('new column for sheet', sheet_name, val1, val2)))

    wb = load_workbook(path)    
    sheet = wb.get_sheet_by_name(sheet_name);

    num_rows = sheet.get_highest_row()
    num_cols = col1 or sheet.get_highest_col()
    start_col = col1 or num_cols
    start_row = row1 or 0
    
    
    cell = sheet.cell(row = start_row, column = start_col)
    cell.value = val1
    for x in range(start_row+1,num_rows):
        cell = sheet.cell(row=x, column=start_col)
        cell.value = val2
    
    wb.save(newpath)   
    print 'wrote new workbook row:', newpath, val1, val2


#     # open our xls file, there's lots of extra default options in this call, 
#     # for logging etc. take a look at the docs
#     book = xlrd.open_workbook(path) 
# 
#     worksheet = None
#     if sheet_name:
#         worksheet = book.sheet_by_name(sheet_name)
#         sheet_index = worksheet.number
#     else:
#         worksheet = book.sheet_by_index(sheet_index) 
#     num_rows = worksheet.nrows
#     num_cols = worksheet.ncols
# 
#     wb = copy(book)
#     worksheet = wb.get_sheet(sheet_index) 
# 
#     start_col = col1 or num_cols
#     start_row = row1 or 0
# 
#     worksheet.write(start_row,start_col,val1)
#     for x in range(start_row+1,num_rows):
#         #worksheet.write(x,start_col,val2)
#         # NOTE: will lose chars in ascii encoding
#         worksheet.write(x,start_col,smart_str(val2, 'ascii', errors='ignore'))    
#     wb.save(newpath)   
#     print 'wrote new workbook col to:', newpath, val1, val2

parser = argparse.ArgumentParser(description='Import file')
parser.add_argument('-f', action='store', dest='inputFile',
                    metavar='FILE', required=True,
                    help='input file path')
parser.add_argument('-of', action='store', dest='outputFile',
                    required=True,
                    help='output file path')
parser.add_argument('-v1', action='store', dest='val1',
                    required=True,
                    help='val 1')
parser.add_argument('-v2', action='store', dest='val2',
                    required=True,
                    help='val 2')
parser.add_argument('-r1', action='store', dest='row1',
                    type=int,
                    help='first row to add a value to')
parser.add_argument('-c1', action='store', dest='col1',
                    type=int,
                    help='first col to add a value to')
parser.add_argument('-v', '--verbose', dest='verbose', action='count',
                help="Increase verbosity (specify multiple times for more)")    
group = parser.add_mutually_exclusive_group()
group.add_argument("-r", "--new_row", action="store_true")
group.add_argument("-c", "--new_col", action="store_true")

parser.add_argument('-sn', '--sheet_name', action='store', required=True,
                    help='sheet name to update')


if __name__ == "__main__":
    args = parser.parse_args()
    if(args.inputFile is None):
        parser.print_help();
        parser.exit(0, "\nMust define the FILE param.\n")

    if not ( args.new_row or args.new_col ):
        parser.print_help()
        parser.exit(0,"\nMust define either the 'new_row' or the 'new_col' param.\n")
         
    log_level = logging.WARNING # default
    if args.verbose == 1:
        log_level = logging.INFO
    elif args.verbose >= 2:
        log_level = logging.DEBUG
    # NOTE this doesn't work because the config is being set by the included 
    # settings.py, and you can only set the config once
    logging.basicConfig(
        level=log_level, 
        format='%(msecs)d:%(module)s:%(lineno)d:%(levelname)s: %(message)s')        
    logger.setLevel(log_level)
        
    print 'run update ', args.inputFile
    
    if args.new_col:
        new_col(args.inputFile, args.outputFile,
                args.val1,args.val2,
                sheet_name=args.sheet_name,
                row1=args.row1, col1=args.col1 )
    elif args.new_row:
        new_row(args.inputFile, args.outputFile,
                args.val1,args.val2,
                sheet_name=args.sheet_name,
                row1=args.row1, col1=args.col1 )
