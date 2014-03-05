import sys
import os
import os.path as op
import re
import codecs
import errno
import shutil
import stat
import argparse
import pandas as pd
import openpyxl
import django.conf
from django.template.loader import render_to_string
import shell_utils as su


# Read RESOURCE_PATH from environ, or default to something that works for me.
RESOURCE_PATH = os.environ.get(
    'RESOURCE_PATH',
    ('/home/jmuhlich/Volumes/imstor on research.files.med.harvard.edu/'
     'sorger/data/NIC/Pat/02-03-2013-videos')
)

PLATEMAP_FILENAME = op.join(RESOURCE_PATH, 'Experiment 02-03-2013.xlsx')
CELL_IMAGE_PREFIX = 'Individual_nolabel_'

LIGAND_ORDER = ['IGF1', 'HRG', 'HGF', 'EGF', 'FGF', 'BTC', 'EPR']
INHIBITOR_ORDER = ['MEKi', 'AKTi', 'MEKi + AKTi']

CONCENTRATION_COLOR_BEGIN = 0xF8
CONCENTRATION_COLOR_END = 0xD0

DOCROOT = op.abspath(op.join(
        op.dirname(__file__), os.pardir,
        'temp', 'docroot', 'explore', 'sampattavanich-2014'
        ))
BASE_URL = '/explore/breast-cancer-signaling/'


def main(argv):

    argparser = argparse.ArgumentParser(
        description='Build sampattavanich_2014 (EKAR) app resources.')
    argparser.add_argument('-n', '--no-media', action='store_true',
                           default=False,
                           help="Don't copy media files)")
    args = argparser.parse_args()

    cur_dir = op.abspath(op.dirname(__file__))
    cell_img_out_dir = op.join(DOCROOT, 'img', 'cell')
    popup_img_out_dir = op.join(DOCROOT, 'img', 'popup')
    movie_out_dir = op.join(DOCROOT, 'movies')

    platemap = build_platemap(PLATEMAP_FILENAME)
    platemap = platemap[platemap.egf_conc == 100]
    inhibitor_concs = [c for c in sorted(platemap.inhibitor_conc.unique())
                       if c > 0]
    inhibitors = platemap.inhibitor.unique()
    batimastat_concs = [c for c in sorted(platemap.batimastat_conc.unique())]

    table = []
    for row, inhibitor_conc in enumerate(inhibitor_concs):
        intensity = (CONCENTRATION_COLOR_BEGIN +
                     (CONCENTRATION_COLOR_END - CONCENTRATION_COLOR_BEGIN) /
                     len(inhibitor_concs) * row)
        color = '#' + ('%02x' % intensity * 3)
        table_row = {'inhibitor_conc': inhibitor_conc,
                     'color': color,
                     'cells': []}
        for batimastat_conc in batimastat_concs:
            for inhibitor in inhibitors:
                location = ((platemap.inhibitor == inhibitor) &
                            (platemap.inhibitor_conc == inhibitor_conc) &
                            (platemap.batimastat_conc == batimastat_conc))
                table_row['cells'].append(platemap[location].iloc[0])
        table.append(table_row)
            
    data = {'inhibitors': inhibitors,
            'table': table,
            'num_inhibitor_concs': len(inhibitor_concs),
            'num_inhibitors': len(inhibitors),
            'batimastat_concs': batimastat_concs,
            'STATIC_URL': django.conf.settings.STATIC_URL,
            'BASE_URL': BASE_URL,
            }
    su.mkdirp(DOCROOT)
    render_template('sampattavanich_2014/ekar/index.html', data,
                    DOCROOT, 'index.html')

    if not args.no_media:
        su.mkdirp(cell_img_out_dir)
        for src_filename in os.listdir(RESOURCE_PATH):
            if not (src_filename.startswith(CELL_IMAGE_PREFIX) and
                    src_filename.endswith('.png')):
                continue
            src_path = op.join(RESOURCE_PATH, src_filename)
            dest_filename = src_filename[len(CELL_IMAGE_PREFIX):]
            dest_path = op.join(cell_img_out_dir, dest_filename)
            shutil.copy(src_path, dest_path)
            os.chmod(dest_path, 0644)
        su.mkdirp(popup_img_out_dir)
        for src_filename in os.listdir(RESOURCE_PATH):
            if not re.match(r'r\d+c\d+\.png', src_filename):
                continue
            src_path = op.join(RESOURCE_PATH, src_filename)
            dest_filename = src_filename
            dest_path = op.join(popup_img_out_dir, dest_filename)
            shutil.copy(src_path, dest_path)
            os.chmod(dest_path, 0644)
        su.mkdirp(movie_out_dir)
        for src_filename in os.listdir(RESOURCE_PATH):
            if not src_filename.endswith('.mp4'):
                continue
            src_path = op.join(RESOURCE_PATH, src_filename)
            dest_filename = re.sub(r'myMov_(\w+)f1ratio', r'\1', src_filename)
            dest_path = op.join(movie_out_dir, dest_filename)
            shutil.copy(src_path, dest_path)
            os.chmod(dest_path, 0644)

    return 0


"""Build plate-map dataframe."""
def build_platemap(filename):

    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[0]

    # Extract row metadata.
    row_meta = dataframe_for_range(ws, 'L3:L10')
    row_meta.columns = pd.Index(['inhibitor'])
    batimastat_conc = float(re.search(r'\d+', ws.cell('M4').value).group())
    egf_conc = float(re.search(r'\d+', ws.cell('N3').value).group())
    row_meta.insert(0, 'plate_row', range(1, len(row_meta)+1))
    row_meta.insert(2, 'batimastat_conc', [0, batimastat_conc] * 4)
    row_meta.insert(3, 'egf_conc', [egf_conc] * 6 + [0] * 2)

    # Extract column metadata.
    col_meta = dataframe_for_range(ws, 'C7:J7').T
    col_meta.columns = pd.Index(['inhibitor_conc'])
    col_meta['inhibitor_conc'] = col_meta['inhibitor_conc'].astype(float)
    first_col = ws.cell('C2').value
    col_meta.insert(0, 'plate_col', range(first_col, first_col+len(col_meta)))

    # Add same-valued dummy columns so merge() will generate a full cartesian
    # product, then delete that column in the resulting dataframe.
    row_meta.insert(0, 'dummy', [0] * len(row_meta))
    col_meta.insert(0, 'dummy', [0] * len(col_meta))
    platemap = pd.merge(row_meta, col_meta, on='dummy')
    del platemap['dummy']

    # Swap the columns around a bit.
    platemap = platemap[[0,4,1,5,2,3]]

    # Replace plate_row and plate_col with a new column in r1c1 format.
    rc_values = platemap.apply(lambda r: 'r%(plate_row)dc%(plate_col)d' % r,
                               axis=1)
    platemap = platemap.drop(['plate_row', 'plate_col'], axis=1)
    platemap.insert(0, 'rc_address', rc_values)

    return platemap


def dataframe_for_range(worksheet, range):
    "Return a Pandas DataFrame from a given range in an openpyxl worksheet."
    data = [[c.value for c in row] for row in worksheet.range(range)]
    return pd.DataFrame(data)


def render_template(template_name, data, dirname, basename):
    "Render a template with data to a file specified by dirname and basename."
    out_filename = op.join(dirname, basename)
    content = render_to_string(template_name, data)
    with codecs.open(out_filename, 'w', 'utf-8') as out_file:
        out_file.write(content)


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
