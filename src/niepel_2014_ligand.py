from __future__ import print_function
from niepel_2014_utils import *
import openpyxl
import requests
import os
import re
import shutil
import argparse
from django.conf import settings


ligand_path = resource_path('SignalingPage', 'LigandPage')
facts_path = resource_path('basic facts_ligands')

image_dirs = [
    ('doseresponse', 'DoseResponsePlots'),
    ('pathwaybias', 'PathwayBiasPlots'),
    ('responsekinetics', 'ResponseKinetics'),
    ('sensitivity', 'SensitivityClass'),
    ('timecourse', 'TimeCoursePlots'),
    ]

image_sizes = {
    'doseresponse': (700, 233),
    'pathwaybias': (700, 292),
    'responsekinetics': (700, 245),
    'sensitivity': (700, 292),
    'timecourse': (700, 233),
}
# Set up dimensions for _large images -- 2x for all.
image_sizes_large = dict((name, (width * 2, height * 2))
                         for (name, (width, height)) in image_sizes.items())

html_path_elements = ['ligand']
img_path_elements = html_path_elements + ['img']
data_path_elements = html_path_elements + ['data']
html_path = create_output_path(*html_path_elements)
data_path = create_output_path(*data_path_elements)

parser = argparse.ArgumentParser(
    description='Build niepel_2014 cell line page resources')
parser.add_argument('-n', '--no-images', action='store_true', default=False,
                    help='Skip building images')
args = parser.parse_args()

print_partial('ligand info')
ligand_info = stash_get('ligand_info')
if not ligand_info:
    filename = os.path.join(facts_path, 'Ligand_info.xlsx')
    workbook = openpyxl.load_workbook(filename, use_iterators=True)
    sheet = workbook.worksheets[0]
    sheet_iter = sheet.iter_rows()
    sheet_iter.next() # skip header row (iter_rows API for offsets is buggy)
    ligand_info = [[cell.value for cell in row] for row in sheet_iter]
    stash_put('ligand_info', ligand_info)
PASS_nl()

print_partial('ligand affinity')
ligand_affinity = stash_get('ligand_affinity')
if not ligand_affinity:
    filename_affinities = os.path.join(facts_path,
                                       'ligand receptor affinities summary.xlsx')
    workbook = openpyxl.load_workbook(filename_affinities, use_iterators=True)
    sheet = workbook.worksheets[0]
    sheet_iter = sheet.iter_rows()
    sheet_iter.next() # skip header row (iter_rows API for offsets is buggy)
    ligand_affinity = [[cell.value for cell in row]
                       for row in sheet_iter]
    stash_put('ligand_affinity', ligand_affinity)
PASS_nl()

column_names = ('hmsl_id', 'family', 'full_name', 'name',
                'vendor_catalog_ignored', 'uniprot_id_ignored', 'concentration',
                'comment')

all_data = []
print()
print(' ' * 15 + 'DR PB RK SE TC RA DB UP  status')

for row in ligand_info:

    data = dict(zip(column_names, row))
    print('%-15s' % data['name'], end='')

    plot_filename = data['name'] + '.png'
    all_ok = all([print_status_accessible(ligand_path, d_in, plot_filename)
                  for d_out, d_in in image_dirs])

    if all_ok:
        for row_affinity in ligand_affinity:
            if row_affinity[0] == data['name']:
                affinity_raw = row_affinity[1:9]
                affinity_pairs = zip(affinity_raw[0::2], affinity_raw[1::2])
                affinity_dicts = [{'name': name, 'kd': kd}
                                  for name, kd in affinity_pairs
                                  if name and kd]
                affinity_reference = row_affinity[9]
                data['affinity'] = {'receptors': affinity_dicts,
                                    'reference': affinity_reference}
                break
        if 'affinity' in data:
            PASS()
        else:
            FAIL()
            all_ok = False

    if all_ok:
        db_stash_key = 'hmsl_db/%s' % data['hmsl_id']
        db_response = stash_get(db_stash_key)
        if not db_response:
            db_url = ('http://lincs.hms.harvard.edu/db/api/v1/protein/%s/' %
                      data['hmsl_id'])
            db_response = requests.get(db_url)
            db_response.raw = None  # hack to make it picklable
            stash_put(db_stash_key, db_response)
        if db_response.ok:
            PASS()
            data['db'] = db_response.json()
        else:
            FAIL()
            all_ok = False

    if all_ok:
        if 'db' in data:
            uniprot_stash_key = 'uniprot/%s' % data['db']['ppUniprotID']
            uniprot_response = stash_get(uniprot_stash_key)
            if not uniprot_response:
                uniprot_url = ('http://www.uniprot.org/uniprot/%s.txt' %
                               data['db']['ppUniprotID'])
                uniprot_response = requests.get(uniprot_url)
                uniprot_response.raw = None  # hack to make it picklable
                stash_put(uniprot_stash_key, uniprot_response)
        # Avoid nested 'if', otherwise we'd need to duplicate the FAIL/all_ok code.
        # Ideally we would write a little more library code to support this idiom.
        if 'db' in data and uniprot_response.ok:
            PASS()
            mw_match = re.search('\nSQ .*?(\d+) MW;', uniprot_response.content)
            molecular_weight = int(mw_match.group(1)) / 1000
            data['molecular_weight'] = molecular_weight
        else:
            FAIL()
            all_ok = False

    if all_ok:
        print(' OK')
        all_data.append(data)
    else:
        print()

stash_put('ligands', all_data)

print()
common = {
    'all_names': [data['name'] for data in all_data],
    'image_sizes': image_sizes,
    'STATIC_URL': settings.STATIC_URL,
    'BASE_URL': BASE_URL,
}
breadcrumb_base = [
    {'url': BASE_URL, 'text': 'Start'},
    {'url': None, 'text': 'Ligands'},
]

for i, data in enumerate(all_data):

    msg = 'rendering page %d/%d %s...' % (i+1, len(all_data), data['name'])
    # FIXME The string padding (37) should be calculated dynamically.
    print_partial('\r' + msg.ljust(37))
    data.update(common)
    data['breadcrumbs'] = breadcrumb_base + [{'url': '', 'text': data['name']}]
    html_filename = data['name'] + '.html'        
    render_template('breast_cancer_signaling/ligand.html', data,
                    html_path, html_filename)
    data_filename = '%s_data.csv' % data['name']
    data_src_path = os.path.join(ligand_path, 'Ligand_data', data_filename)
    data_dest_path = os.path.join(data_path, data_filename)
    shutil.copy(data_src_path, data_dest_path)

    if args.no_images:
        continue
    image_filename = data['name'] + '.png'
    copy_images(image_dirs, image_filename, ligand_path, img_path_elements,
                new_sizes=image_sizes, new_format='jpg',
                format_options={'quality': 85, 'optimize': True})
    (if_root, if_ext) = os.path.splitext(image_filename)
    if_root + '_large' + if_ext
    copy_images(image_dirs, image_filename, ligand_path, img_path_elements,
                new_sizes=image_sizes_large, new_format='jpg',
                dest_suffix='_large',
                format_options={'quality': 75, 'optimize': True})

print_partial("done")
PASS_nl()
