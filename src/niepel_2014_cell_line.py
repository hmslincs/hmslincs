from __future__ import print_function
from niepel_2014_utils import *
import openpyxl
import requests
import os
import re
import shutil
import argparse
from django.conf import settings


cellline_path = resource_path('SignalingPage', 'CellLinePage')
facts_path = resource_path('basic facts_cell lines')

image_dirs = [
    ('foldchange', 'FoldChangeBoxPlot'),
    ('nodeedge', 'NodeEdgeFigures'),
    ('subtype', 'subtypeBoxPlot'),
    ('topmeasures', 'BasalTopMeasures'),
    ]

image_sizes = {
    'foldchange': [700, 237],
    'nodeedge': [250, 235],
    'subtype': [700, 237],
    'topmeasures': [700, 385],
    }
# Set up dimensions for _large images - start with x2 for all images.
image_sizes_large = dict((name, [width * 2, height * 2])
                         for name, [width, height] in image_sizes.items())
# Override with x3 for nodeedge.
image_sizes_large['nodeedge'] = [dim*3 for dim in image_sizes['nodeedge']]
# Override height for topmeasures, which uses a wholly different image for the
# large size.
image_sizes_large['topmeasures'][1] = 1867

topmeasures_large_suffix = '_allRTKs'

html_path_elements = ['cell-line']
img_path_elements = html_path_elements + ['img']
data_path_elements = html_path_elements + ['data']
html_path = create_output_path(*html_path_elements)
data_path = create_output_path(*data_path_elements)

parser = argparse.ArgumentParser(
    description='Build niepel_2014 cell line page resources')
parser.add_argument('-n', '--no-images', action='store_true', default=False,
                    help='Skip building images')
args = parser.parse_args()

print_partial('cell line info')
cellline_filename = os.path.join(facts_path, 'CellLine_subset_BMCpaper.xlsx')
mutation_filename = os.path.join(facts_path, 'CellLine_PTEN_PI3K_mutations.xlsx')
dest_cellline_filename = os.path.join(data_path, 'cell_line_info.xlsx')
dest_mutation_filename = os.path.join(data_path, 'cell_line_pten_pi3k_mutations.xlsx')
shutil.copy(cellline_filename, dest_cellline_filename)
shutil.copy(mutation_filename, dest_mutation_filename)
cellline_info = stash_get('cellline_info')
if not cellline_info:
    workbook = openpyxl.load_workbook(cellline_filename, use_iterators=True)
    sheet = workbook.worksheets[0]
    sheet_iter = sheet.iter_rows()
    sheet_iter.next() # skip header row (iter_rows API for offsets is buggy)
    cellline_info = [[cell.value for cell in row] for row in sheet_iter]
    stash_put('cellline_info', cellline_info)
PASS_nl()

column_names = ('hmsl_id', 'name', 'short_name', 'is_icbp43', 'atcc_id_ignored',
                'vendor', 'class_gagdar', 'class_neve_oe', 'class_neve_gc',
                'class_kao_rs', 'class_kao_tr', 'class_heiser', 'notes',
                'class_niepel', 'growth_medium', 'culture_temperature',
                'culture_atmosphere')

all_data = []
print()
print(' ' * 15 + 'FC NE ST TM DB  status')

for row in cellline_info:

    data = dict(zip(column_names, row))
    print('%-15s' % data['name'], end='')

    plot_filename = data['name'] + '.png'
    all_ok = all([print_status_accessible(cellline_path, d_in, plot_filename)
                  for d_out, d_in in image_dirs])

    if all_ok:
        db_stash_key = 'hmsl_db/%s' % data['hmsl_id']
        db_response = stash_get(db_stash_key)
        if not db_response:
            db_url = 'http://lincs.hms.harvard.edu/db/api/v1/cell/%s/' % data['hmsl_id']
            db_response = requests.get(db_url)
            db_response.raw = None  # hack to make it picklable
            stash_put(db_stash_key, db_response)
        if db_response.ok:
            PASS()
            data['db'] = db_response.json()
            if data['db']['clAlternateID']:
                cosmic_match = re.search(r'COSMIC:\s*(\d+)', data['db']['clAlternateID'])
                if cosmic_match:
                    data['db']['cosmic_id'] = cosmic_match.group(1)
        else:
            FAIL()
            all_ok = False

    if all_ok:
        print(' OK')
        all_data.append(data)
    else:
        print()

stash_put('cell_lines', all_data)

print()
common = {
    'all_names': [data['name'] for data in all_data],
    'image_sizes': image_sizes,
    'STATIC_URL': settings.STATIC_URL,
    'BASE_URL': BASE_URL,
}
breadcrumb_base = [
    {'url': BASE_URL, 'text': 'Start'},
    {'url': None, 'text': 'Cell lines'},
]

for i, data in enumerate(all_data):

    msg = 'rendering page %d/%d %s...' % (i+1, len(all_data), data['name'])
    # FIXME The string padding (37) should be calculated dynamically.
    print_partial('\r' + msg.ljust(37))
    data.update(common)
    data['breadcrumbs'] = breadcrumb_base + [{'url': '', 'text': data['name']}]
    html_filename = data['name'] + '.html'
    render_template('breast_cancer_signaling/cell_line.html', data,
                    html_path, html_filename)
    data_filename = '%s_data.csv' % data['name']
    data_src_path = os.path.join(cellline_path, 'CellLine_data', data_filename)
    data_dest_path = os.path.join(data_path, data_filename)
    shutil.copy(data_src_path, data_dest_path)

    if args.no_images:
        continue
    image_filename = data['name'] + '.png'
    copy_images(image_dirs, image_filename,
                cellline_path, img_path_elements,
                new_sizes=image_sizes, new_format='jpg',
                format_options={'quality': 85, 'optimize': True})
    (if_root, if_ext) = os.path.splitext(image_filename)
    if_root + '_large' + if_ext
    image_dirs_no_topmeasures = [d for d in image_dirs if d[0] != 'topmeasures']
    copy_images(image_dirs_no_topmeasures, image_filename,
                cellline_path, img_path_elements,
                new_sizes=image_sizes_large, new_format='jpg',
                dest_suffix='_large',
                format_options={'quality': 75, 'optimize': True})
    image_dirs_topmeasures = [d for d in image_dirs if d[0] == 'topmeasures']
    if_root, if_ext = os.path.splitext(image_filename)
    image_filename_topmeasures = if_root + topmeasures_large_suffix + if_ext
    copy_images(image_dirs_topmeasures, image_filename_topmeasures,
                cellline_path, img_path_elements,
                new_sizes=image_sizes_large, new_format='jpg',
                dest_suffix='_large',
                format_options={'quality': 75, 'optimize': True})

print_partial("done")
PASS_nl()

if not args.no_images:
    print()
    subtypes = set(d['class_niepel'] for d in all_data)
    for subtype in subtypes:
        image_filename = 'NetMap_%s.png' % subtype
        # We are only copying one image, but we can reuse copy_images with a
        # little creativity in crafting the first arg.
        copy_images([('nodeedge', 'NodeEdgeFigures')], image_filename,
                    cellline_path, img_path_elements,
                    new_sizes=image_sizes, new_format='jpg',
                    format_options={'quality': 85, 'optimize': True})
        copy_images([('nodeedge', 'NodeEdgeFigures')], image_filename,
                    cellline_path, img_path_elements,
                    new_sizes=image_sizes_large, new_format='jpg',
                    dest_suffix='_large',
                    format_options={'quality': 85, 'optimize': True})
        print(image_filename, end=' ')
        PASS_nl()
