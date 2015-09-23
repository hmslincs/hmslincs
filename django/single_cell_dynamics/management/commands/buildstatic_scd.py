import re
import collections
from optparse import make_option
import unipath
import openpyxl
import numpy as np
import pandas as pd
import h5py
from django.template.loader import render_to_string
from django.core.management.base import BaseCommand, CommandError
from django.contrib.flatpages.models import FlatPage
from django.contrib.sites.models import Site
from hmslincs.management_utils import SimpleLoggingMixin


class Command(BaseCommand, SimpleLoggingMixin):
    help = 'Builds the static assets and html chunks'
    option_list = BaseCommand.option_list + (
        make_option('-r', '--resource-path',
                    default=('/home/jmuhlich/Volumes/'
                             'imstor on files.med.harvard.edu/sorger/data/'
                             'computation/Jeremy/single_cell_dynamics'),
                    help=('Path to resource files (contains '
                          '"Experiment 02-03-2013.xlsx")')),
        make_option('-b', '--no-big-data', action='store_true', default=False,
                    help="Don't copy big zipped data file (2.5GB)"),
        make_option('-d', '--no-data', action='store_true', default=False,
                    help="Don't copy data files"),
        make_option('-m', '--no-media', action='store_true', default=False,
                    help="Don't copy media files"),
    )

    def handle(self, *args, **options):

        self.configure_logging(options)

        url = '/explore/single-cell-dynamics/'
        content = self.build_static(options)

        page, created = FlatPage.objects.get_or_create(url=url)
        page.title = ('Dynamics of perturbagen responses in living single '
                      'cells')
        page.content = content
        page.template_name = 'single_cell_dynamics/base.html'
        page.sites.clear()
        page.sites.add(Site.objects.get_current())
        page.save()


    def build_static(self, options):

        app_path = unipath.Path(__file__).absolute().ancestor(3)
        static_path = app_path.child('static', 'single_cell_dynamics')
        generated_path = static_path.child('g')
        data_out_dir = generated_path.child('data')
        img_out_dir = generated_path.child('img')
        cell_img_out_dir = img_out_dir.child('cell')
        popup_img_out_dir = img_out_dir.child('popup')
        movie_out_dir = generated_path.child('movies')

        resource_path = unipath.Path(options['resource_path'])
        platemap_filename = resource_path.child('Experiment 02-03-2013.xlsx')
        data_in_dir = resource_path.child('data')
        cell_img_in_dir = resource_path.child('img', 'cell')
        popup_img_in_dir = resource_path.child('img', 'popup')
        movie_in_dir = resource_path.child('movies')

        ligand_order = ['IGF1', 'HRG', 'HGF', 'EGF', 'FGF', 'BTC', 'EPR']
        inhibitor_order = ['MEKi', 'AKTi', 'MEKi + AKTi']

        concentration_color_begin = 0xF8
        concentration_color_end = 0xD0

        platemap = build_platemap(platemap_filename)
        platemap = platemap[platemap.egf_conc == 100]
        inhibitor_concs = [c for c in sorted(platemap.inhibitor_conc.unique())
                           if c > 0]
        inhibitors = platemap.inhibitor.unique()
        batimastat_concs = [c for c in sorted(platemap.batimastat_conc.unique())]

        table = []
        for row, inhibitor_conc in enumerate(inhibitor_concs):
            intensity = (concentration_color_begin +
                         (concentration_color_end - concentration_color_begin) /
                         len(inhibitor_concs) * row)
            color = '#' + ('%02x' % intensity * 3)
            table_row = {'inhibitor_conc': inhibitor_conc,
                         'color': color,
                         'cells': []}
            for batimastat_conc in batimastat_concs:
                for inhibitor in inhibitors:
                    location = ((platemap.inhibitor == inhibitor) &
                                (platemap.inhibitor_conc == inhibitor_conc) &
                                (platemap.batimastat_conc == batimastat_conc))
                    table_row['cells'].append(platemap[location].iloc[0])
            table.append(table_row)

        # Assemble data for template and render html.
        data = {
            'inhibitors': inhibitors,
            'table': table,
            'num_inhibitor_concs': len(inhibitor_concs),
            'num_inhibitors': len(inhibitors),
            'batimastat_concs': batimastat_concs,
        }
        content = render_to_string('single_cell_dynamics/index.html', data)

        generated_path.mkdir()

        if not options['no_big_data']:
            self.info('Copying big zipped data file:')
            data_out_dir.mkdir()
            src_path = data_in_dir.child('HMS-LINCS-ekarev-processed-data.zip')
            dest_path = data_out_dir.child(src_path.name)
            self.info('  %s', src_path.name)
            src_path.copy(dest_path)
            dest_path.chmod(0644)
        else:
            self.info('Skipping big zipped data file')

        if not options['no_data']:
            self.info('Converting data files:')
            data_out_dir.mkdir()
            for src_path in data_in_dir.listdir():
                match = re.match(r'H5OUT_(r\d+)_(c\d+)\.h5', src_path.name)
                if not match:
                    self.debug('  ignoring %s', src_path.name)
                    continue
                dest_filename = 'ekar_single_cell_data_%s%s.csv' % match.groups()
                dest_path = data_out_dir.child(dest_filename)
                self.info('  %s -> %s', src_path.name, dest_filename)
                data = extract_well_data(src_path)
                with open(dest_path, 'w') as f:
                    data.to_csv(f, index=False)
        else:
            self.info('Skipping data files')

        if not options['no_media']:
            img_out_dir.mkdir()
            self.info('Copying cell image files:')
            cell_img_out_dir.mkdir()
            for src_path in cell_img_in_dir.listdir():
                if not re.match(r'r\d+c\d+\.png', src_path.name):
                    self.debug('  ignoring %s', src_path.name)
                    continue
                dest_path = cell_img_out_dir.child(src_path.name)
                self.info('  %s', src_path.name)
                src_path.copy(dest_path)
                dest_path.chmod(0644)
            self.info('Copying popup image files:')
            popup_img_out_dir.mkdir()
            for src_path in popup_img_in_dir.listdir():
                if not re.match(r'r\d+c\d+\.png', src_path.name):
                    self.debug('  ignoring %s', src_path.name)
                    continue
                dest_path = popup_img_out_dir.child(src_path.name)
                self.info('  %s', src_path.name)
                src_path.copy(dest_path)
                dest_path.chmod(0644)
            self.info('Copying movie files:')
            movie_out_dir.mkdir()
            for src_path in movie_in_dir.listdir():
                if src_path.ext != '.mp4':
                    self.debug('  ignoring %s', src_path.name)
                    continue
                dest_path = movie_out_dir.child(src_path.name)
                self.info('  %s', src_path.name)
                src_path.copy(dest_path)
                dest_path.chmod(0644)
        else:
            self.info('Skipping media files')

        return content


"""Build plate-map dataframe."""
def build_platemap(filename):

    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[0]

    # Extract row metadata.
    row_meta = dataframe_for_range(ws, 'L3:L10')
    row_meta.columns = pd.Index(['inhibitor'])
    batimastat_conc = float(re.search(r'\d+', ws.cell('M4').value).group())
    egf_conc = float(re.search(r'\d+', ws.cell('N3').value).group())
    row_meta.insert(0, 'plate_row', range(1, len(row_meta)+1))
    row_meta.insert(2, 'batimastat_conc', [0, batimastat_conc] * 4)
    row_meta.insert(3, 'egf_conc', [egf_conc] * 6 + [0] * 2)

    # Extract column metadata.
    col_meta = dataframe_for_range(ws, 'C7:J7').T
    col_meta.columns = pd.Index(['inhibitor_conc'])
    col_meta['inhibitor_conc'] = col_meta['inhibitor_conc'].astype(float)
    first_col = ws.cell('C2').value
    col_meta.insert(0, 'plate_col', range(first_col, first_col+len(col_meta)))

    # Add same-valued dummy columns so merge() will generate a full cartesian
    # product, then delete that column in the resulting dataframe.
    row_meta.insert(0, 'dummy', [0] * len(row_meta))
    col_meta.insert(0, 'dummy', [0] * len(col_meta))
    platemap = pd.merge(row_meta, col_meta, on='dummy')
    del platemap['dummy']

    # Swap the columns around a bit.
    platemap = platemap[[0,4,1,5,2,3]]

    # Replace plate_row and plate_col with a new column in r1c1 format.
    rc_values = platemap.apply(lambda r: 'r%(plate_row)dc%(plate_col)d' % r,
                               axis=1)
    platemap = platemap.drop(['plate_row', 'plate_col'], axis=1)
    platemap.insert(0, 'rc_address', rc_values)

    return platemap


def dataframe_for_range(worksheet, range):
    "Return a Pandas DataFrame from a given range in an openpyxl worksheet."
    data = [[c.value for c in row] for row in worksheet.iter_rows(range)]
    return pd.DataFrame(data)


def extract_well_data(filename):
    "Extract dataframe from one per-well HDF5 file."
    f = h5py.File(filename, 'r')
    field1 = f['field1']
    num_cells = field1['outputsignal1'].shape[2]
    num_times = field1['timestamp1'].shape[0]
    # Here we collect the various data columns, for all cells and all
    # timepoints. The rows are to be ordered first by cell, then by timepoint.
    cell_id = np.repeat(np.arange(0, num_cells), num_times)
    # Index 1 in outputsignal1 is "signal2" ("Cytosol-Math").
    erk_signal = field1['outputsignal1'][1, :].T.flatten()
    x = field1['cellpath'][:,0,:].T.flatten()
    y = field1['cellpath'][:,1,:].T.flatten()
    time = np.tile(field1['timestamp1'][:], num_cells)
    data = collections.OrderedDict([
            ('cell_id', cell_id), ('time', time), ('x', x), ('y', y),
            ('erk_signal', erk_signal)])
    df = pd.DataFrame(data)
    return df
